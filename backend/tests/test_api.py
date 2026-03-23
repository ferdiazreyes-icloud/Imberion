"""Basic API tests for USG Pricing Decision Engine."""

from datetime import date

from app.models import Territory, Category, Product, Customer, Transaction, Elasticity, Recommendation


def _seed_basic_data(db):
    """Insert minimal test data."""
    territory = Territory(region="Norte", state="Nuevo León", municipality="Monterrey")
    db.add(territory)
    db.flush()

    customer = Customer(name="Test Distributor", type="distribuidor", segment="oro", territory_id=territory.id)
    db.add(customer)
    db.flush()

    category = Category(name="Tableros")
    db.add(category)
    db.flush()

    product = Product(sku_code="TB-001", name="Tabla Roca Regular 1/2\"", category_id=category.id)
    db.add(product)
    db.flush()

    transaction = Transaction(
        date=date(2024, 6, 1),
        customer_id=customer.id,
        product_id=product.id,
        territory_id=territory.id,
        volume=100.0,
        list_price=250.0,
        discount=20.0,
        rebate=10.0,
        net_price=220.0,
        revenue=22000.0,
    )
    db.add(transaction)

    elasticity = Elasticity(
        type="historical",
        coefficient=-1.2,
        confidence_level="high",
        p_value=0.01,
        r_squared=0.85,
        node_type="sku",
        node_id=product.id,
        period_start=date(2024, 1, 1),
        period_end=date(2025, 12, 31),
        sample_size=500,
    )
    db.add(elasticity)

    recommendation = Recommendation(
        product_id=product.id,
        segment="oro",
        territory_id=territory.id,
        action_type="increase",
        suggested_change_pct=5.0,
        expected_impact_revenue=100000.0,
        expected_impact_volume=-500.0,
        expected_impact_margin=30000.0,
        confidence_level="high",
        rationale={"elasticity": -1.2},
    )
    db.add(recommendation)
    db.commit()

    return {
        "territory": territory,
        "customer": customer,
        "category": category,
        "product": product,
    }


def test_health(client):
    response = client.get("/api/health")
    assert response.status_code == 200
    assert response.json()["status"] == "ok"


def test_filters_segments(client):
    response = client.get("/api/filters/segments")
    assert response.status_code == 200
    assert response.json() == ["oro", "plata", "bronce"]


def test_filters_regions(client, db):
    _seed_basic_data(db)
    response = client.get("/api/filters/regions")
    assert response.status_code == 200
    assert "Norte" in response.json()


def test_filters_categories(client, db):
    _seed_basic_data(db)
    response = client.get("/api/filters/categories")
    assert response.status_code == 200
    assert len(response.json()) >= 1
    assert response.json()[0]["name"] == "Tableros"


def test_overview(client, db):
    _seed_basic_data(db)
    response = client.get("/api/overview")
    assert response.status_code == 200
    data = response.json()
    assert "total_revenue" in data
    assert "total_volume" in data


def test_history_elasticities(client, db):
    _seed_basic_data(db)
    response = client.get("/api/history/elasticities")
    assert response.status_code == 200
    data = response.json()
    assert len(data) >= 1
    assert data[0]["coefficient"] == -1.2


def test_recommendations(client, db):
    _seed_basic_data(db)
    response = client.get("/api/recommendations")
    assert response.status_code == 200
    data = response.json()
    assert len(data) >= 1
    assert data[0]["action_type"] == "increase"
    assert data[0]["product_name"] == "Tabla Roca Regular 1/2\""


def test_recommendations_summary(client, db):
    _seed_basic_data(db)
    response = client.get("/api/recommendations/summary")
    assert response.status_code == 200
    assert len(response.json()) >= 1


def test_quick_simulate(client, db):
    _seed_basic_data(db)
    response = client.get("/api/simulator/quick-simulate?product_id=1&price_change_pct=5")
    assert response.status_code == 200
    data = response.json()
    assert "curve" in data
    assert len(data["curve"]) == 21  # -20 to +20 in steps of 2


def test_passthrough_by_segment(client, db):
    _seed_basic_data(db)
    response = client.get("/api/passthrough/by-segment")
    assert response.status_code == 200


def test_export_csv(client, db):
    _seed_basic_data(db)
    response = client.get("/api/export/recommendations-csv")
    assert response.status_code == 200
    assert "text/csv" in response.headers["content-type"]


def test_export_executive_summary(client, db):
    _seed_basic_data(db)
    response = client.get("/api/export/executive-summary")
    assert response.status_code == 200
    data = response.json()
    assert "title" in data
    assert "top_recommendations" in data


# ---------------------------------------------------------------------------
# Enhanced simulator tests
# ---------------------------------------------------------------------------

def _create_test_scenario(client, db):
    """Seed data and create a scenario, return scenario id."""
    _seed_basic_data(db)
    resp = client.post("/api/simulator/scenarios", json={
        "name": "Test +5%",
        "description": "Test scenario",
        "price_changes": [{"change_pct": 5.0}],
    })
    assert resp.status_code == 200
    return resp.json()["id"]


def test_create_scenario_uses_real_margin(client, db):
    """Verify that create_scenario uses predict_scenario (not flat 30%)."""
    sid = _create_test_scenario(client, db)
    resp = client.get(f"/api/simulator/scenarios/{sid}/results")
    assert resp.status_code == 200
    results = resp.json()
    assert len(results) >= 1
    r = results[0]
    # With predict_scenario (cost_pct=0.70), margin should NOT equal revenue * 0.30
    # For a +5% change with elasticity -1.2:
    #   new_volume = 100 * (1 + (-1.2 * 0.05)) = 94
    #   new_price = 220 * 1.05 = 231
    #   new_revenue = 94 * 231 = 21714
    #   unit_cost = 220 * 0.70 = 154
    #   new_margin = (231 - 154) * 94 = 7238
    # So margin should be ~7238, NOT 21714 * 0.30 = 6514.2
    assert r["expected_margin"] > 0
    assert r["expected_volume"] > 0
    # Margin should NOT be exactly 30% of revenue (old behavior)
    ratio = r["expected_margin"] / r["expected_revenue"] if r["expected_revenue"] else 0
    assert abs(ratio - 0.30) > 0.001, "Margin should use real cost calculation, not flat 30%"


def test_quick_simulate_uses_real_margin(client, db):
    """Verify quick-simulate curve uses predict_scenario with real cost calc.
    At +10% price, margin/revenue ratio should differ from 30% because
    unit cost stays fixed while price increases."""
    _seed_basic_data(db)
    resp = client.get("/api/simulator/quick-simulate?product_id=1&price_change_pct=5")
    data = resp.json()
    # At +10% change, margin should NOT equal revenue * 0.30
    # because unit_cost = base_price * 0.70 stays fixed while new_price increases
    point_10 = next(p for p in data["curve"] if p["price_change_pct"] == 10)
    ratio = point_10["margin"] / point_10["revenue"] if point_10["revenue"] else 0
    assert ratio > 0.30, "At +10% price, margin ratio should exceed 30% (fixed cost model)"


def test_scenario_summary(client, db):
    sid = _create_test_scenario(client, db)
    resp = client.get(f"/api/simulator/scenarios/{sid}/summary")
    assert resp.status_code == 200
    data = resp.json()
    assert data["scenario_name"] == "Test +5%"
    assert "total_volume" in data
    assert "total_revenue" in data
    assert "total_margin" in data
    assert "delta_volume" in data
    assert "delta_revenue_pct" in data
    assert len(data["by_category"]) >= 1
    assert data["by_category"][0]["name"] == "Tableros"


def test_scenario_results_grouped_by_category(client, db):
    sid = _create_test_scenario(client, db)
    resp = client.get(f"/api/simulator/scenarios/{sid}/results-grouped?group_by=category")
    assert resp.status_code == 200
    data = resp.json()
    assert len(data) >= 1
    assert data[0]["group_name"] == "Tableros"
    assert data[0]["product_count"] >= 1
    assert "total_revenue" in data[0]
    assert "avg_confidence" in data[0]


def test_scenario_results_grouped_by_segment(client, db):
    """Create scenario with segment and group by segment."""
    _seed_basic_data(db)
    resp = client.post("/api/simulator/scenarios", json={
        "name": "Segment test",
        "price_changes": [{"change_pct": 3.0, "segment": "oro"}],
    })
    sid = resp.json()["id"]
    resp = client.get(f"/api/simulator/scenarios/{sid}/results-grouped?group_by=segment")
    assert resp.status_code == 200


def test_compare_multi_scenarios(client, db):
    _seed_basic_data(db)
    # Create two scenarios
    r1 = client.post("/api/simulator/scenarios", json={
        "name": "Low +3%", "price_changes": [{"change_pct": 3.0}],
    })
    r2 = client.post("/api/simulator/scenarios", json={
        "name": "High +10%", "price_changes": [{"change_pct": 10.0}],
    })
    sid1 = r1.json()["id"]
    sid2 = r2.json()["id"]

    resp = client.get(f"/api/simulator/compare-multi?scenario_ids={sid1},{sid2}")
    assert resp.status_code == 200
    data = resp.json()
    assert len(data["scenarios"]) == 2
    assert "rankings" in data
    assert "best_for_volume" in data["rankings"]
    assert "best_for_margin" in data["rankings"]
    # The +3% scenario should have more volume (less elastic loss)
    assert data["rankings"]["best_for_volume"] == sid1


def test_compare_multi_requires_two(client):
    resp = client.get("/api/simulator/compare-multi?scenario_ids=1")
    assert resp.status_code == 400


def test_best_scenario(client, db):
    _seed_basic_data(db)
    client.post("/api/simulator/scenarios", json={
        "name": "A", "price_changes": [{"change_pct": 5.0}],
    })
    client.post("/api/simulator/scenarios", json={
        "name": "B", "price_changes": [{"change_pct": -5.0}],
    })

    resp = client.get("/api/simulator/best-scenario?objective=margin")
    assert resp.status_code == 200
    data = resp.json()
    assert data["objective"] == "margin"
    assert "best" in data
    assert data["best"]["scenario_name"] in ["A", "B"]
    assert "runners_up" in data


def test_best_scenario_no_scenarios(client, db):
    """Should return 404 when no scenarios exist."""
    _seed_basic_data(db)
    resp = client.get("/api/simulator/best-scenario?objective=volume")
    assert resp.status_code == 404


# ---------------------------------------------------------------------------
# customer_id filter tests
# ---------------------------------------------------------------------------

def test_overview_with_customer_id(client, db):
    data = _seed_basic_data(db)
    resp = client.get(f"/api/overview?customer_id={data['customer'].id}")
    assert resp.status_code == 200
    assert resp.json()["total_volume"]["value"] > 0


def test_overview_with_invalid_customer_id(client, db):
    _seed_basic_data(db)
    resp = client.get("/api/overview?customer_id=9999")
    assert resp.status_code == 200
    assert resp.json()["total_volume"]["value"] == 0


def test_passthrough_with_customer_id(client, db):
    data = _seed_basic_data(db)
    resp = client.get(f"/api/passthrough/by-segment?customer_id={data['customer'].id}")
    assert resp.status_code == 200


def test_history_trends_with_customer_id(client, db):
    data = _seed_basic_data(db)
    resp = client.get(f"/api/history/trends?customer_id={data['customer'].id}")
    assert resp.status_code == 200


def test_filters_customers(client, db):
    _seed_basic_data(db)
    resp = client.get("/api/filters/customers")
    assert resp.status_code == 200
    customers = resp.json()
    assert len(customers) >= 1
    assert customers[0]["name"] == "Test Distributor"


def test_filters_customers_by_segment(client, db):
    _seed_basic_data(db)
    resp = client.get("/api/filters/customers?segment=oro")
    assert resp.status_code == 200
    assert len(resp.json()) >= 1
    resp2 = client.get("/api/filters/customers?segment=plata")
    assert resp2.status_code == 200
    assert len(resp2.json()) == 0


# ---------------------------------------------------------------------------
# Scenario CSV export test
# ---------------------------------------------------------------------------

def test_export_scenario_csv(client, db):
    sid = _create_test_scenario(client, db)
    resp = client.get(f"/api/export/scenario-csv/{sid}")
    assert resp.status_code == 200
    assert "text/csv" in resp.headers["content-type"]
    content = resp.content.decode("utf-8")
    assert "Product" in content
    assert "Expected Volume" in content


def test_export_scenario_csv_not_found(client, db):
    _seed_basic_data(db)
    resp = client.get("/api/export/scenario-csv/9999")
    assert resp.status_code == 404


# ---------------------------------------------------------------------------
# Multi-ID (comma-separated) filter tests
# ---------------------------------------------------------------------------

def test_overview_with_comma_separated_customer_ids(client, db):
    """Comma-separated customer_id should filter to those customers."""
    data = _seed_basic_data(db)
    cid = data["customer"].id
    # Single ID as string
    resp = client.get(f"/api/overview?customer_id={cid}")
    assert resp.status_code == 200
    assert resp.json()["total_volume"]["value"] > 0
    # Comma-separated with a nonexistent ID — should still return data for the valid one
    resp2 = client.get(f"/api/overview?customer_id={cid},9999")
    assert resp2.status_code == 200
    assert resp2.json()["total_volume"]["value"] > 0


def test_passthrough_with_comma_separated_territory_ids(client, db):
    data = _seed_basic_data(db)
    tid = data["territory"].id
    resp = client.get(f"/api/passthrough/by-segment?territory_id={tid},9999")
    assert resp.status_code == 200


# ---------------------------------------------------------------------------
# Excel template & upload tests
# ---------------------------------------------------------------------------

def test_download_template_excel(client):
    resp = client.get("/api/simulator/template-excel")
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]


def test_upload_excel_scenario(client, db):
    """Upload an Excel with one SKU price change and verify scenario creation."""
    import openpyxl
    import io

    _seed_basic_data(db)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["sku_code", "category", "change_pct", "segment", "territory"])
    ws.append(["TB-001", "", 5.0, "oro", ""])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = client.post(
        "/api/simulator/scenarios/from-excel",
        files={"file": ("test.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"name": "Excel Test", "objective": "margin"},
    )
    assert resp.status_code == 200
    data = resp.json()
    assert data["parsed_rows"] == 1
    assert data["scenario"]["name"] == "Excel Test"
    assert "suggestions" in data


def test_upload_excel_invalid_sku(client, db):
    """Excel with unknown SKU should report errors."""
    import openpyxl
    import io

    _seed_basic_data(db)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["sku_code", "category", "change_pct", "segment", "territory"])
    ws.append(["INVALID-SKU", "", 5.0, "", ""])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = client.post(
        "/api/simulator/scenarios/from-excel",
        files={"file": ("test.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"name": "Bad Excel"},
    )
    assert resp.status_code == 400
    assert "No valid rows" in resp.json()["detail"]


def test_upload_excel_by_category(client, db):
    """Excel with category-level price change."""
    import openpyxl
    import io

    _seed_basic_data(db)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["sku_code", "category", "change_pct", "segment", "territory"])
    ws.append(["", "Tableros", 3.0, "", ""])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = client.post(
        "/api/simulator/scenarios/from-excel",
        files={"file": ("test.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"name": "Category Excel"},
    )
    assert resp.status_code == 200
    assert resp.json()["parsed_rows"] == 1


# ---------------------------------------------------------------------------
# Optimization endpoint tests
# ---------------------------------------------------------------------------

def test_optimize_scenario(client, db):
    """Optimize for margin with default range."""
    _seed_basic_data(db)
    resp = client.post("/api/simulator/scenarios/optimize", json={
        "name": "Optimize Margin",
        "objective": "margin",
        "price_min_pct": -10.0,
        "price_max_pct": 15.0,
    })
    assert resp.status_code == 200
    data = resp.json()
    assert data["objective"] == "margin"
    assert data["products_optimized"] >= 1
    assert data["scenario"]["name"] == "Optimize Margin"
    assert len(data["optimization_details"]) >= 1


def test_optimize_scenario_volume(client, db):
    """Optimize for volume."""
    _seed_basic_data(db)
    resp = client.post("/api/simulator/scenarios/optimize", json={
        "name": "Optimize Volume",
        "objective": "volume",
        "price_min_pct": -20.0,
        "price_max_pct": 0.0,
    })
    assert resp.status_code == 200
    data = resp.json()
    assert data["objective"] == "volume"
    # For volume maximization with negative elasticity, optimal should be a price decrease
    detail = data["optimization_details"][0]
    assert detail["optimal_change_pct"] <= 0


def test_optimize_scenario_with_category_filter(client, db):
    """Optimize with category filter."""
    data = _seed_basic_data(db)
    resp = client.post("/api/simulator/scenarios/optimize", json={
        "name": "Optimize Category",
        "objective": "revenue",
        "price_min_pct": -5.0,
        "price_max_pct": 10.0,
        "category_id": str(data["category"].id),
    })
    assert resp.status_code == 200
    assert resp.json()["products_optimized"] >= 1
