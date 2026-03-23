from typing import List, Optional
import io

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.database import get_db
from app.models import (
    Transaction, Customer, Product, Territory, Category,
    Elasticity, Scenario, ScenarioResult,
)
from app.schemas.scenario import (
    ScenarioCreate, ScenarioOut, ScenarioResultOut,
    ScenarioSummaryResponse, GroupedResultOut,
    MultiCompareResponse, ScenarioCompareItem, ScenarioRankings,
    BestScenarioResponse, OptimizeRequest, SuggestionItem, ExcelScenarioResponse,
)
from app.analytics.prediction_model import predict_scenario, optimal_price_search, suggest_improvements
from app.analytics.confidence_scorer import score_confidence

router = APIRouter()


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _compute_base_totals(db: Session, product_ids: list[int]) -> dict:
    """Return base volume/revenue totals keyed by product_id."""
    if not product_ids:
        return {}
    rows = (
        db.query(
            Transaction.product_id,
            func.avg(Transaction.net_price).label("price"),
            func.sum(Transaction.volume).label("volume"),
            func.sum(Transaction.revenue).label("revenue"),
        )
        .filter(Transaction.product_id.in_(product_ids))
        .group_by(Transaction.product_id)
        .all()
    )
    return {r.product_id: r for r in rows}


def _scenario_totals(db: Session, scenario_id: int) -> dict:
    """Compute aggregate totals for a scenario and its base."""
    results = db.query(ScenarioResult).filter(ScenarioResult.scenario_id == scenario_id).all()
    if not results:
        return {"total_volume": 0, "total_revenue": 0, "total_margin": 0,
                "base_volume": 0, "base_revenue": 0, "base_margin": 0}

    product_ids = [r.product_id for r in results]
    base_totals = _compute_base_totals(db, product_ids)

    total_volume = sum(r.expected_volume for r in results)
    total_revenue = sum(r.expected_revenue for r in results)
    total_margin = sum(r.expected_margin for r in results)

    base_volume = sum(float(bt.volume or 0) for bt in base_totals.values())
    base_revenue = sum(float(bt.revenue or 0) for bt in base_totals.values())
    base_margin = base_revenue * 0.30  # base margin approximation

    return {
        "total_volume": round(total_volume, 2),
        "total_revenue": round(total_revenue, 2),
        "total_margin": round(total_margin, 2),
        "base_volume": round(base_volume, 2),
        "base_revenue": round(base_revenue, 2),
        "base_margin": round(base_margin, 2),
    }


@router.get("/simulator/scenarios", response_model=List[ScenarioOut])
def list_scenarios(db: Session = Depends(get_db)):
    return db.query(Scenario).order_by(Scenario.created_at.desc()).all()


@router.post("/simulator/scenarios", response_model=ScenarioOut)
def create_scenario(data: ScenarioCreate, db: Session = Depends(get_db)):
    scenario = Scenario(
        name=data.name,
        description=data.description,
        assumptions=data.assumptions or {},
    )
    db.add(scenario)
    db.flush()

    for pc in data.price_changes:
        product_ids = []
        if pc.product_id:
            product_ids = [pc.product_id]
        elif pc.category_id:
            product_ids = [p.id for p in db.query(Product.id).filter(Product.category_id == pc.category_id).all()]
        else:
            product_ids = [p.id for p in db.query(Product.id).all()]

        # Batch fetch elasticities for all products
        elasticities = {
            e.node_id: e
            for e in db.query(Elasticity).filter(
                Elasticity.node_type == "sku",
                Elasticity.node_id.in_(product_ids),
            ).all()
        }

        # Batch fetch base data for all products
        base_q = (
            db.query(
                Transaction.product_id,
                func.avg(Transaction.net_price).label("price"),
                func.sum(Transaction.volume).label("volume"),
                func.sum(Transaction.revenue).label("revenue"),
            )
            .filter(Transaction.product_id.in_(product_ids))
            .group_by(Transaction.product_id)
        )
        if pc.segment:
            base_q = base_q.join(Customer).filter(Customer.segment == pc.segment)
        if pc.territory_id:
            base_q = base_q.filter(Transaction.territory_id == pc.territory_id)

        base_data = {row.product_id: row for row in base_q.all()}

        for pid in product_ids:
            # Get best elasticity: prefer predicted, fallback to historical
            elast = elasticities.get(pid)
            if elast and elast.type != "predicted":
                predicted = next(
                    (e for e in elasticities.values() if e.node_id == pid and e.type == "predicted"), None
                )
                if predicted:
                    elast = predicted

            coefficient = elast.coefficient if elast else -1.0

            # Use confidence scorer instead of static assignment
            if elast:
                conf = score_confidence(elast.p_value, elast.r_squared, elast.sample_size)
                confidence = conf["confidence_level"]
            else:
                confidence = "low"

            base = base_data.get(pid)
            base_price = float(base.price) if base and base.price else 100.0
            base_volume = float(base.volume) if base and base.volume else 0.0

            # Use predict_scenario for real margin calculation
            pred = predict_scenario(
                base_price=base_price,
                base_volume=base_volume,
                price_change_pct=pc.change_pct,
                elasticity=coefficient,
            )

            result = ScenarioResult(
                scenario_id=scenario.id,
                product_id=pid,
                segment=pc.segment,
                territory_id=pc.territory_id,
                price_change_pct=pc.change_pct,
                expected_volume=pred["new_volume"],
                expected_revenue=pred["new_revenue"],
                expected_margin=pred["new_margin"],
                confidence_level=confidence,
            )
            db.add(result)

    db.commit()
    db.refresh(scenario)
    return scenario


@router.get("/simulator/scenarios/{scenario_id}/results", response_model=List[ScenarioResultOut])
def get_scenario_results(scenario_id: int, db: Session = Depends(get_db)):
    rows = (
        db.query(ScenarioResult, Product, Territory)
        .join(Product, ScenarioResult.product_id == Product.id)
        .outerjoin(Territory, ScenarioResult.territory_id == Territory.id)
        .filter(ScenarioResult.scenario_id == scenario_id)
        .all()
    )
    return [
        ScenarioResultOut(
            id=r.id,
            scenario_id=r.scenario_id,
            product_id=r.product_id,
            product_name=p.name,
            segment=r.segment,
            territory_name=t.state if t else None,
            price_change_pct=r.price_change_pct,
            expected_volume=r.expected_volume,
            expected_revenue=r.expected_revenue,
            expected_margin=r.expected_margin,
            confidence_level=r.confidence_level,
        )
        for r, p, t in rows
    ]


@router.get("/simulator/compare")
def compare_scenarios(scenario_id: int, db: Session = Depends(get_db)):
    """Compare a scenario against actual transaction data."""
    scenario = db.get(Scenario, scenario_id)
    if not scenario:
        raise HTTPException(status_code=404, detail="Scenario not found")

    scenario_results = db.query(ScenarioResult).filter(ScenarioResult.scenario_id == scenario_id).all()

    # Batch fetch base data
    product_ids = [r.product_id for r in scenario_results]
    base_totals = {}
    if product_ids:
        base_rows = (
            db.query(
                Transaction.product_id,
                func.sum(Transaction.volume).label("volume"),
                func.sum(Transaction.revenue).label("revenue"),
            )
            .filter(Transaction.product_id.in_(product_ids))
            .group_by(Transaction.product_id)
            .all()
        )
        base_totals = {row.product_id: row for row in base_rows}

    base_data = []
    for r in scenario_results:
        bt = base_totals.get(r.product_id)
        base_data.append({
            "product_id": r.product_id,
            "segment": r.segment,
            "base_volume": float(bt.volume) if bt and bt.volume else 0.0,
            "base_revenue": float(bt.revenue) if bt and bt.revenue else 0.0,
            "scenario_volume": r.expected_volume,
            "scenario_revenue": r.expected_revenue,
            "scenario_margin": r.expected_margin,
            "price_change_pct": r.price_change_pct,
            "confidence_level": r.confidence_level,
        })

    total_base_revenue = sum(d["base_revenue"] for d in base_data)
    total_scenario_revenue = sum(d["scenario_revenue"] for d in base_data)
    total_base_volume = sum(d["base_volume"] for d in base_data)
    total_scenario_volume = sum(d["scenario_volume"] for d in base_data)
    total_scenario_margin = sum(d["scenario_margin"] for d in base_data)

    return {
        "scenario_name": scenario.name,
        "details": base_data,
        "delta_revenue": round(total_scenario_revenue - total_base_revenue, 2),
        "delta_volume": round(total_scenario_volume - total_base_volume, 2),
        "total_margin": round(total_scenario_margin, 2),
    }


@router.get("/simulator/quick-simulate")
def quick_simulate(
    product_id: Optional[int] = None,
    category_id: Optional[str] = None,
    segment: Optional[str] = None,
    customer_id: Optional[str] = None,
    price_change_pct: float = 5.0,
    db: Session = Depends(get_db),
):
    """Quick simulation without persisting. Returns price-volume-margin curve."""
    from app.api.overview import _parse_ids, _parse_strs, _filter_ids

    cat_ids = _parse_ids(category_id)
    # Fetch elasticity and base data ONCE outside the loop
    if product_id:
        elast = db.query(Elasticity).filter(
            Elasticity.node_type == "sku", Elasticity.node_id == product_id
        ).first()
        base_q = db.query(
            func.avg(Transaction.net_price), func.sum(Transaction.volume), func.sum(Transaction.revenue)
        ).filter(Transaction.product_id == product_id)
    elif cat_ids:
        elast = db.query(Elasticity).filter(
            Elasticity.node_type == "category", Elasticity.node_id == cat_ids[0]
        ).first()
        base_q = db.query(
            func.avg(Transaction.net_price), func.sum(Transaction.volume), func.sum(Transaction.revenue)
        ).join(Product).filter(Product.category_id.in_(cat_ids) if len(cat_ids) > 1 else Product.category_id == cat_ids[0])
    else:
        elast = db.query(Elasticity).filter(Elasticity.node_type == "category").first()
        base_q = db.query(
            func.avg(Transaction.net_price), func.sum(Transaction.volume), func.sum(Transaction.revenue)
        )

    base_q = _filter_ids(base_q, Transaction.customer_id, _parse_ids(customer_id))
    segs = _parse_strs(segment)
    if segs:
        base_q = base_q.join(Customer, Customer.id == Transaction.customer_id).filter(Customer.segment.in_(segs) if len(segs) > 1 else Customer.segment == segs[0])

    row = base_q.one_or_none()
    if not row or row[0] is None:
        return {"elasticity_used": -1.0, "confidence": "low", "curve": []}
    base_price = float(row[0] or 100)
    base_volume = float(row[1] or 0)
    coefficient = elast.coefficient if elast else -1.0

    # Use confidence scorer for robust confidence level
    if elast:
        conf = score_confidence(elast.p_value, elast.r_squared, elast.sample_size)
        confidence = conf["confidence_level"]
    else:
        confidence = "low"

    # Generate curve using predict_scenario for real margin calculation
    points = []
    for pct in range(-20, 21, 2):
        pred = predict_scenario(
            base_price=base_price,
            base_volume=base_volume,
            price_change_pct=float(pct),
            elasticity=coefficient,
        )
        points.append({
            "price_change_pct": pct,
            "price": pred["new_price"],
            "volume": pred["new_volume"],
            "revenue": pred["new_revenue"],
            "margin": pred["new_margin"],
        })

    return {
        "elasticity_used": coefficient,
        "confidence": confidence,
        "curve": points,
    }


# ---------------------------------------------------------------------------
# Endpoint 2: Scenario summary with aggregates
# ---------------------------------------------------------------------------

@router.get("/simulator/scenarios/{scenario_id}/summary", response_model=ScenarioSummaryResponse)
def get_scenario_summary(scenario_id: int, db: Session = Depends(get_db)):
    """Aggregated summary of a scenario with breakdowns by category and segment."""
    scenario = db.get(Scenario, scenario_id)
    if not scenario:
        raise HTTPException(status_code=404, detail="Scenario not found")

    totals = _scenario_totals(db, scenario_id)

    # Breakdown by category
    cat_rows = (
        db.query(
            Category.id,
            Category.name,
            func.sum(ScenarioResult.expected_volume).label("vol"),
            func.sum(ScenarioResult.expected_revenue).label("rev"),
            func.sum(ScenarioResult.expected_margin).label("mar"),
            func.count(ScenarioResult.id).label("cnt"),
        )
        .join(Product, Product.id == ScenarioResult.product_id)
        .join(Category, Category.id == Product.category_id)
        .filter(ScenarioResult.scenario_id == scenario_id)
        .group_by(Category.id, Category.name)
        .all()
    )
    by_category = [
        {"name": r.name, "id": r.id, "total_volume": round(float(r.vol), 2),
         "total_revenue": round(float(r.rev), 2), "total_margin": round(float(r.mar), 2),
         "product_count": r.cnt}
        for r in cat_rows
    ]

    # Breakdown by segment
    seg_rows = (
        db.query(
            ScenarioResult.segment,
            func.sum(ScenarioResult.expected_volume).label("vol"),
            func.sum(ScenarioResult.expected_revenue).label("rev"),
            func.sum(ScenarioResult.expected_margin).label("mar"),
            func.count(ScenarioResult.id).label("cnt"),
        )
        .filter(ScenarioResult.scenario_id == scenario_id)
        .group_by(ScenarioResult.segment)
        .all()
    )
    by_segment = [
        {"name": r.segment or "Sin segmento", "total_volume": round(float(r.vol), 2),
         "total_revenue": round(float(r.rev), 2), "total_margin": round(float(r.mar), 2),
         "product_count": r.cnt}
        for r in seg_rows
    ]

    base_rev = totals["base_revenue"]
    base_vol = totals["base_volume"]
    base_mar = totals["base_margin"]

    return ScenarioSummaryResponse(
        scenario_name=scenario.name,
        total_volume=totals["total_volume"],
        total_revenue=totals["total_revenue"],
        total_margin=totals["total_margin"],
        base_volume=base_vol,
        base_revenue=base_rev,
        base_margin=base_mar,
        delta_volume=round(totals["total_volume"] - base_vol, 2),
        delta_volume_pct=round((totals["total_volume"] / base_vol - 1) * 100, 2) if base_vol else 0,
        delta_revenue=round(totals["total_revenue"] - base_rev, 2),
        delta_revenue_pct=round((totals["total_revenue"] / base_rev - 1) * 100, 2) if base_rev else 0,
        delta_margin=round(totals["total_margin"] - base_mar, 2),
        delta_margin_pct=round((totals["total_margin"] / base_mar - 1) * 100, 2) if base_mar else 0,
        by_category=by_category,
        by_segment=by_segment,
    )


# ---------------------------------------------------------------------------
# Endpoint 3: Drill-down results grouped by level
# ---------------------------------------------------------------------------

@router.get("/simulator/scenarios/{scenario_id}/results-grouped", response_model=List[GroupedResultOut])
def get_scenario_results_grouped(
    scenario_id: int,
    group_by: str = Query("category", pattern="^(category|segment|territory)$"),
    db: Session = Depends(get_db),
):
    """Return scenario results aggregated by category, segment, or territory."""
    scenario = db.get(Scenario, scenario_id)
    if not scenario:
        raise HTTPException(status_code=404, detail="Scenario not found")

    if group_by == "category":
        rows = (
            db.query(
                Category.id.label("gid"),
                Category.name.label("gname"),
                func.sum(ScenarioResult.expected_volume).label("vol"),
                func.sum(ScenarioResult.expected_revenue).label("rev"),
                func.sum(ScenarioResult.expected_margin).label("mar"),
                func.count(ScenarioResult.id).label("cnt"),
                func.avg(ScenarioResult.price_change_pct).label("avg_pct"),
            )
            .join(Product, Product.id == ScenarioResult.product_id)
            .join(Category, Category.id == Product.category_id)
            .filter(ScenarioResult.scenario_id == scenario_id)
            .group_by(Category.id, Category.name)
            .all()
        )
    elif group_by == "segment":
        rows = (
            db.query(
                ScenarioResult.segment.label("gid"),
                ScenarioResult.segment.label("gname"),
                func.sum(ScenarioResult.expected_volume).label("vol"),
                func.sum(ScenarioResult.expected_revenue).label("rev"),
                func.sum(ScenarioResult.expected_margin).label("mar"),
                func.count(ScenarioResult.id).label("cnt"),
                func.avg(ScenarioResult.price_change_pct).label("avg_pct"),
            )
            .filter(ScenarioResult.scenario_id == scenario_id)
            .group_by(ScenarioResult.segment)
            .all()
        )
    else:  # territory
        rows = (
            db.query(
                Territory.id.label("gid"),
                Territory.state.label("gname"),
                func.sum(ScenarioResult.expected_volume).label("vol"),
                func.sum(ScenarioResult.expected_revenue).label("rev"),
                func.sum(ScenarioResult.expected_margin).label("mar"),
                func.count(ScenarioResult.id).label("cnt"),
                func.avg(ScenarioResult.price_change_pct).label("avg_pct"),
            )
            .join(Territory, Territory.id == ScenarioResult.territory_id)
            .filter(ScenarioResult.scenario_id == scenario_id)
            .group_by(Territory.id, Territory.state)
            .all()
        )

    # Determine majority confidence per group
    def _majority_confidence(scenario_id: int, group_by: str, gid) -> str:
        q = db.query(ScenarioResult.confidence_level, func.count().label("c")).filter(
            ScenarioResult.scenario_id == scenario_id
        )
        if group_by == "category":
            q = q.join(Product, Product.id == ScenarioResult.product_id).filter(Product.category_id == gid)
        elif group_by == "segment":
            q = q.filter(ScenarioResult.segment == gid)
        else:
            q = q.filter(ScenarioResult.territory_id == gid)
        conf_rows = q.group_by(ScenarioResult.confidence_level).order_by(func.count().desc()).first()
        return conf_rows[0] if conf_rows else "low"

    return [
        GroupedResultOut(
            group_key=str(r.gid),
            group_name=r.gname or "N/A",
            total_volume=round(float(r.vol), 2),
            total_revenue=round(float(r.rev), 2),
            total_margin=round(float(r.mar), 2),
            product_count=r.cnt,
            avg_price_change_pct=round(float(r.avg_pct), 2),
            avg_confidence=_majority_confidence(scenario_id, group_by, r.gid),
        )
        for r in rows
    ]


# ---------------------------------------------------------------------------
# Endpoint 4: Multi-scenario comparison
# ---------------------------------------------------------------------------

@router.get("/simulator/compare-multi", response_model=MultiCompareResponse)
def compare_multi_scenarios(
    scenario_ids: str = Query(..., description="Comma-separated scenario IDs"),
    db: Session = Depends(get_db),
):
    """Compare multiple scenarios side-by-side with rankings."""
    ids = [int(x.strip()) for x in scenario_ids.split(",") if x.strip()]
    if len(ids) < 2:
        raise HTTPException(status_code=400, detail="At least 2 scenario IDs required")

    items = []
    for sid in ids:
        scenario = db.get(Scenario, sid)
        if not scenario:
            raise HTTPException(status_code=404, detail=f"Scenario {sid} not found")

        totals = _scenario_totals(db, sid)
        base_rev = totals["base_revenue"]
        base_vol = totals["base_volume"]
        base_mar = totals["base_margin"]

        items.append(ScenarioCompareItem(
            scenario_id=sid,
            scenario_name=scenario.name,
            total_volume=totals["total_volume"],
            total_revenue=totals["total_revenue"],
            total_margin=totals["total_margin"],
            delta_volume=round(totals["total_volume"] - base_vol, 2),
            delta_revenue=round(totals["total_revenue"] - base_rev, 2),
            delta_margin=round(totals["total_margin"] - base_mar, 2),
            delta_volume_pct=round((totals["total_volume"] / base_vol - 1) * 100, 2) if base_vol else 0,
            delta_revenue_pct=round((totals["total_revenue"] / base_rev - 1) * 100, 2) if base_rev else 0,
            delta_margin_pct=round((totals["total_margin"] / base_mar - 1) * 100, 2) if base_mar else 0,
        ))

    rankings = ScenarioRankings(
        best_for_volume=max(items, key=lambda x: x.total_volume).scenario_id,
        best_for_revenue=max(items, key=lambda x: x.total_revenue).scenario_id,
        best_for_margin=max(items, key=lambda x: x.total_margin).scenario_id,
    )

    return MultiCompareResponse(scenarios=items, rankings=rankings)


# ---------------------------------------------------------------------------
# Endpoint 5: Best scenario recommendation
# ---------------------------------------------------------------------------

@router.get("/simulator/best-scenario", response_model=BestScenarioResponse)
def get_best_scenario(
    objective: str = Query("margin", pattern="^(margin|volume|revenue)$"),
    db: Session = Depends(get_db),
):
    """Find the best saved scenario for a given objective."""
    scenarios = db.query(Scenario).filter(Scenario.is_base == False).all()  # noqa: E712
    if not scenarios:
        raise HTTPException(status_code=404, detail="No scenarios saved yet")

    items = []
    for s in scenarios:
        totals = _scenario_totals(db, s.id)
        base_rev = totals["base_revenue"]
        base_vol = totals["base_volume"]
        base_mar = totals["base_margin"]

        items.append(ScenarioCompareItem(
            scenario_id=s.id,
            scenario_name=s.name,
            total_volume=totals["total_volume"],
            total_revenue=totals["total_revenue"],
            total_margin=totals["total_margin"],
            delta_volume=round(totals["total_volume"] - base_vol, 2),
            delta_revenue=round(totals["total_revenue"] - base_rev, 2),
            delta_margin=round(totals["total_margin"] - base_mar, 2),
            delta_volume_pct=round((totals["total_volume"] / base_vol - 1) * 100, 2) if base_vol else 0,
            delta_revenue_pct=round((totals["total_revenue"] / base_rev - 1) * 100, 2) if base_rev else 0,
            delta_margin_pct=round((totals["total_margin"] / base_mar - 1) * 100, 2) if base_mar else 0,
        ))

    key_map = {"margin": "total_margin", "volume": "total_volume", "revenue": "total_revenue"}
    sorted_items = sorted(items, key=lambda x: getattr(x, key_map[objective]), reverse=True)

    return BestScenarioResponse(
        objective=objective,
        best=sorted_items[0],
        runners_up=sorted_items[1:4],
    )


# ---------------------------------------------------------------------------
# Shared helper: create scenario results from price changes
# ---------------------------------------------------------------------------

def _create_scenario_results(db: Session, scenario_id: int, price_changes: list) -> list:
    """
    Process a list of PriceChange-like dicts and create ScenarioResults.
    Each item: {product_id?, category_id?, segment?, territory_id?, change_pct}
    Returns list of created ScenarioResult objects.
    """
    results = []
    for pc in price_changes:
        product_ids = []
        pid = pc.get("product_id")
        cid = pc.get("category_id")
        if pid:
            product_ids = [pid]
        elif cid:
            product_ids = [p.id for p in db.query(Product.id).filter(Product.category_id == cid).all()]
        else:
            product_ids = [p.id for p in db.query(Product.id).all()]

        elasticities = {
            e.node_id: e
            for e in db.query(Elasticity).filter(
                Elasticity.node_type == "sku",
                Elasticity.node_id.in_(product_ids),
            ).all()
        }

        base_q = (
            db.query(
                Transaction.product_id,
                func.avg(Transaction.net_price).label("price"),
                func.sum(Transaction.volume).label("volume"),
                func.sum(Transaction.revenue).label("revenue"),
            )
            .filter(Transaction.product_id.in_(product_ids))
            .group_by(Transaction.product_id)
        )
        seg = pc.get("segment")
        tid = pc.get("territory_id")
        if seg:
            base_q = base_q.join(Customer).filter(Customer.segment == seg)
        if tid:
            base_q = base_q.filter(Transaction.territory_id == tid)

        base_data = {row.product_id: row for row in base_q.all()}

        for p_id in product_ids:
            elast = elasticities.get(p_id)
            if elast and elast.type != "predicted":
                predicted = next(
                    (e for e in elasticities.values() if e.node_id == p_id and e.type == "predicted"), None
                )
                if predicted:
                    elast = predicted

            coefficient = elast.coefficient if elast else -1.0
            if elast:
                conf = score_confidence(elast.p_value, elast.r_squared, elast.sample_size)
                confidence = conf["confidence_level"]
            else:
                confidence = "low"

            base = base_data.get(p_id)
            base_price = float(base.price) if base and base.price else 100.0
            base_volume = float(base.volume) if base and base.volume else 0.0

            pred = predict_scenario(
                base_price=base_price,
                base_volume=base_volume,
                price_change_pct=pc["change_pct"],
                elasticity=coefficient,
            )

            result = ScenarioResult(
                scenario_id=scenario_id,
                product_id=p_id,
                segment=seg,
                territory_id=tid,
                price_change_pct=pc["change_pct"],
                expected_volume=pred["new_volume"],
                expected_revenue=pred["new_revenue"],
                expected_margin=pred["new_margin"],
                confidence_level=confidence,
            )
            db.add(result)
            results.append(result)

    return results


# ---------------------------------------------------------------------------
# Endpoint 6: Excel template download
# ---------------------------------------------------------------------------

@router.get("/simulator/template-excel")
def download_template_excel():
    """Download Excel template for scenario upload."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Price Changes"

    headers = ["sku_code", "category", "change_pct", "segment", "territory"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = openpyxl.styles.Font(bold=True)

    # Example rows
    examples = [
        ("TB-001", "", 5.0, "oro", ""),
        ("", "Plafones", -3.0, "", "Nuevo León"),
        ("PL-005", "", 2.5, "plata", "CDMX"),
    ]
    for i, row in enumerate(examples, 2):
        for col, val in enumerate(row, 1):
            ws.cell(row=i, column=col, value=val)

    # Auto-width
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col) + 2
        ws.column_dimensions[col[0].column_letter].width = max_len

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=scenario_template.xlsx"},
    )


# ---------------------------------------------------------------------------
# Endpoint 7: Excel upload → create scenario + suggestions
# ---------------------------------------------------------------------------

@router.post("/simulator/scenarios/from-excel")
def create_scenario_from_excel(
    file: UploadFile = File(...),
    name: str = Form("Excel Scenario"),
    objective: str = Form("margin"),
    db: Session = Depends(get_db),
):
    """Upload Excel with price changes, create scenario, and get improvement suggestions."""
    import openpyxl

    # Parse Excel
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file.file.read()))
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid Excel file")

    # Build lookups
    sku_to_product = {p.sku_code: p for p in db.query(Product).all()}
    cat_name_to_id = {c.name.lower(): c.id for c in db.query(Category).all()}
    state_to_territory = {t.state.lower(): t.id for t in db.query(Territory).all()}

    price_changes = []
    errors = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(v is None for v in row):
            continue

        sku_code = str(row[0]).strip() if row[0] else ""
        category = str(row[1]).strip() if row[1] else ""
        change_pct = row[2]
        segment = str(row[3]).strip().lower() if row[3] else None
        territory = str(row[4]).strip().lower() if row[4] else None

        if change_pct is None:
            errors.append(f"Row {row_idx}: missing change_pct")
            continue

        try:
            change_pct = float(change_pct)
        except (ValueError, TypeError):
            errors.append(f"Row {row_idx}: invalid change_pct '{change_pct}'")
            continue

        pc = {"change_pct": change_pct}

        if sku_code and sku_code in sku_to_product:
            pc["product_id"] = sku_to_product[sku_code].id
        elif category and category.lower() in cat_name_to_id:
            pc["category_id"] = cat_name_to_id[category.lower()]
        elif sku_code:
            errors.append(f"Row {row_idx}: unknown SKU '{sku_code}'")
            continue
        elif category:
            errors.append(f"Row {row_idx}: unknown category '{category}'")
            continue
        else:
            errors.append(f"Row {row_idx}: must specify sku_code or category")
            continue

        if segment and segment in ("oro", "plata", "bronce"):
            pc["segment"] = segment
        if territory and territory in state_to_territory:
            pc["territory_id"] = state_to_territory[territory]

        price_changes.append(pc)

    if not price_changes:
        raise HTTPException(status_code=400, detail=f"No valid rows found. Errors: {errors}")

    # Create scenario
    scenario = Scenario(
        name=name,
        assumptions={"source": "excel", "errors": errors, "rows_parsed": len(price_changes)},
    )
    db.add(scenario)
    db.flush()

    _create_scenario_results(db, scenario.id, price_changes)
    db.commit()
    db.refresh(scenario)

    # Generate suggestions
    planned = []
    base_dict = {}
    elast_dict = {}
    for pc in price_changes:
        pid = pc.get("product_id")
        if not pid:
            continue
        planned.append({"product_id": pid, "change_pct": pc["change_pct"]})

    if planned:
        pids = [p["product_id"] for p in planned]
        base_rows = (
            db.query(
                Transaction.product_id,
                func.avg(Transaction.net_price).label("price"),
                func.sum(Transaction.volume).label("volume"),
            )
            .filter(Transaction.product_id.in_(pids))
            .group_by(Transaction.product_id)
            .all()
        )
        base_dict = {r.product_id: {"price": float(r.price or 100), "volume": float(r.volume or 0)} for r in base_rows}

        elast_rows = db.query(Elasticity).filter(
            Elasticity.node_type == "sku", Elasticity.node_id.in_(pids)
        ).all()
        elast_dict = {e.node_id: e.coefficient for e in elast_rows}

    raw_suggestions = suggest_improvements(planned, base_dict, elast_dict, objective=objective)

    # Enrich suggestions with product names
    product_names = {p.id: p.name for p in db.query(Product).all()}
    suggestions = [
        SuggestionItem(
            product_name=product_names.get(s["product_id"], ""),
            **s,
        )
        for s in raw_suggestions
    ]

    return {
        "scenario": ScenarioOut.model_validate(scenario),
        "parsed_rows": len(price_changes),
        "errors": errors,
        "suggestions": suggestions,
    }


# ---------------------------------------------------------------------------
# Endpoint 8: Automatic optimization
# ---------------------------------------------------------------------------

@router.post("/simulator/scenarios/optimize")
def create_optimized_scenario(
    data: OptimizeRequest,
    db: Session = Depends(get_db),
):
    """Create a scenario with automatically optimized prices."""
    from app.api.overview import _parse_ids, _parse_strs

    # Determine product scope
    product_q = db.query(Product)
    cat_ids = _parse_ids(data.category_id)
    if cat_ids:
        product_q = product_q.filter(Product.category_id.in_(cat_ids))
    products = product_q.all()
    product_ids = [p.id for p in products]

    if not product_ids:
        raise HTTPException(status_code=400, detail="No products in scope")

    # Fetch elasticities
    elasticities = {
        e.node_id: e
        for e in db.query(Elasticity).filter(
            Elasticity.node_type == "sku",
            Elasticity.node_id.in_(product_ids),
        ).all()
    }

    # Fetch base data (optionally filtered by segment/territory/customer)
    base_q = (
        db.query(
            Transaction.product_id,
            func.avg(Transaction.net_price).label("price"),
            func.sum(Transaction.volume).label("volume"),
        )
        .filter(Transaction.product_id.in_(product_ids))
        .group_by(Transaction.product_id)
    )
    segs = _parse_strs(data.segment)
    if segs:
        base_q = base_q.join(Customer).filter(Customer.segment.in_(segs))
    tid_list = _parse_ids(data.territory_id)
    if tid_list:
        base_q = base_q.filter(Transaction.territory_id.in_(tid_list))
    cust_ids = _parse_ids(data.customer_id)
    if cust_ids:
        base_q = base_q.filter(Transaction.customer_id.in_(cust_ids))

    base_data = {r.product_id: r for r in base_q.all()}

    # Run optimization per product
    price_changes = []
    optimization_details = []
    for pid in product_ids:
        elast = elasticities.get(pid)
        coefficient = elast.coefficient if elast else -1.0

        base = base_data.get(pid)
        base_price = float(base.price) if base and base.price else 100.0
        base_volume = float(base.volume) if base and base.volume else 0.0

        if base_volume == 0:
            continue

        opt = optimal_price_search(
            base_price=base_price,
            base_volume=base_volume,
            elasticity=coefficient,
            objective=data.objective,
            price_range=(data.price_min_pct, data.price_max_pct),
        )

        price_changes.append({
            "product_id": pid,
            "change_pct": opt["optimal_change_pct"],
            "segment": data.segment,
            "territory_id": tid_list[0] if tid_list and len(tid_list) == 1 else None,
        })

        optimization_details.append({
            "product_id": pid,
            "optimal_change_pct": opt["optimal_change_pct"],
            "optimal_value": opt["optimal_value"],
        })

    if not price_changes:
        raise HTTPException(status_code=400, detail="No products with transaction data in scope")

    # Create scenario
    scenario = Scenario(
        name=data.name,
        assumptions={
            "source": "optimization",
            "objective": data.objective,
            "price_range": [data.price_min_pct, data.price_max_pct],
            "filters": {
                "segment": data.segment,
                "territory_id": data.territory_id,
                "customer_id": data.customer_id,
                "category_id": data.category_id,
            },
        },
    )
    db.add(scenario)
    db.flush()

    _create_scenario_results(db, scenario.id, price_changes)
    db.commit()
    db.refresh(scenario)

    return {
        "scenario": ScenarioOut.model_validate(scenario),
        "products_optimized": len(price_changes),
        "objective": data.objective,
        "price_range": [data.price_min_pct, data.price_max_pct],
        "optimization_details": optimization_details,
    }
